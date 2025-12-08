from __future__ import annotations

from typing import Dict, List, Mapping, Tuple
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


class ExcelReadError(Exception):
    """Custom exception used when reading Excel fails."""
    pass


TableRangeConfig = Mapping[str, List[str]]
TableColumns = Dict[str, List[str]]  # {"Employees": ["Id", "Name", ...]}


def split_coord(coord: str) -> Tuple[str, int]:
    """Split a cell reference like 'A1' or 'BC12' into (letters, row_number)."""
    coord = coord.strip()
    if not coord:
        raise ValueError("Empty cell reference")

    letters = ""
    digits = ""

    for ch in coord:
        if ch.isalpha():
            letters += ch.upper()
        elif ch.isdigit():
            digits += ch
        else:
            raise ValueError(f"Invalid character {ch!r} in cell reference {coord!r}")

    if not letters or not digits:
        raise ValueError(f"Invalid cell reference {coord!r}")

    return letters, int(digits)


def read_excel_ranges(
    filepath: str,
    tables_config: TableRangeConfig,
    sheet_name: str | None = None,
) -> TableColumns:
    """Read an Excel file and return columns for each table."""

    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as exc:
        raise ExcelReadError(f"Failed to open file '{filepath}': {exc}") from exc

    try:
        ws = wb[sheet_name] if sheet_name else wb.active
    except KeyError as exc:
        wb.close()
        raise ExcelReadError(f"Sheet '{sheet_name}' not found in '{filepath}'.") from exc

    result: TableColumns = {}

    for key, value in tables_config.items():
        if not isinstance(value, (list, tuple)) or len(value) != 3:
            continue

        user_table_name, start_coord, end_coord = value
        user_table_name = (user_table_name or "").strip() or key

        try:
            start_letters, start_row = split_coord(start_coord)
            end_letters, end_row = split_coord(end_coord)

            min_col = column_index_from_string(start_letters)
            max_col = column_index_from_string(end_letters)

            min_col, max_col = sorted((min_col, max_col))
            header_row = min(start_row, end_row)

        except Exception as exc:
            wb.close()
            raise ExcelReadError(
                f"Invalid coordinates for table '{user_table_name}': "
                f"{start_coord!r} - {end_coord!r}. Error: {exc}"
            ) from exc

        headers: List[str] = []
        for col_cells in ws.iter_cols(
            min_col=min_col,
            max_col=max_col,
            min_row=header_row,
            max_row=header_row,
        ):
            cell = col_cells[0]
            if cell.value is not None:
                headers.append(str(cell.value))

        result[user_table_name] = headers

    wb.close()
    return result


def read_excel_for_controller(controller, sheet_name: str | None = None) -> TableColumns:
    filepath = getattr(controller, "filepath", None)
    tables_config = getattr(controller, "tables_data", None)

    if not filepath:
        raise ExcelReadError("controller.filepath is not set.")
    if not tables_config:
        raise ExcelReadError("controller.tables_data is empty or not set.")

    return read_excel_ranges(filepath, tables_config, sheet_name=sheet_name)


# ----------------------------------------------------------
# STANDALONE TEST MODE (RUN WITHOUT GUI)
# ----------------------------------------------------------

if __name__ == "__main__":
    from pprint import pprint

    # The Excel file is in the same folder as this script
    filepath = "2025-inkomna-ansokningar-lan-kommun.xlsx"

    # We read the first sheet: "Lista ansökningar"
    # Header row = row 1
    # Columns A → O (15 columns)
    tables_config: TableRangeConfig = {
        "Table 1": ["Ansökningar", "A1", "O1"],
    }

    try:
        result = read_excel_ranges(
            filepath=filepath,
            tables_config=tables_config,
            sheet_name="Lista ansökningar",
        )
    except ExcelReadError as e:
        print("ExcelReadError:", e)
    else:
        print("\n✅ Headers found:")
        pprint(result)
